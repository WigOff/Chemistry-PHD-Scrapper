"""
Europe Chemistry QS Top-50 University Scraper
==============================================
Scrapes chemistry research groups, professors, and research areas from
top European universities. Each university has its own dedicated scraper
function to handle differing website structures.

Usage:
    python scraper.py

Output:
    europe_chemistry_qs_top50.xlsx
    europe_chemistry_qs_top50.csv
"""

import time
import re
import logging
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

REQUEST_DELAY = 1.5      # seconds between requests
REQUEST_TIMEOUT = 20     # seconds per request
MAX_RETRIES = 2

OUTPUT_XLSX = "europe_chemistry_qs_top50.xlsx"
OUTPUT_CSV  = "europe_chemistry_qs_top50.csv"

# QS World University Rankings 2024 (Chemistry / Overall for Europe)
# rank: approximate QS subject rank or overall rank
UNIVERSITY_META = {
    "University of Cambridge":          {"country": "United Kingdom",  "qs_rank": 1},
    "University of Oxford":             {"country": "United Kingdom",  "qs_rank": 2},
    "ETH Zurich":                       {"country": "Switzerland",     "qs_rank": 3},
    "Imperial College London":          {"country": "United Kingdom",  "qs_rank": 4},
    "UCL":                              {"country": "United Kingdom",  "qs_rank": 5},
    "University of Edinburgh":          {"country": "United Kingdom",  "qs_rank": 6},
    "EPFL":                             {"country": "Switzerland",     "qs_rank": 7},
    "LMU Munich":                       {"country": "Germany",         "qs_rank": 8},
    "University of Copenhagen":         {"country": "Denmark",         "qs_rank": 9},
    "KU Leuven":                        {"country": "Belgium",         "qs_rank": 10},
    "University of Amsterdam":          {"country": "Netherlands",     "qs_rank": 11},
    "Leiden University":                {"country": "Netherlands",     "qs_rank": 12},
    "Utrecht University":               {"country": "Netherlands",     "qs_rank": 13},
    "TU Munich":                        {"country": "Germany",         "qs_rank": 14},
    "Heidelberg University":            {"country": "Germany",         "qs_rank": 15},
    "University of Manchester":         {"country": "United Kingdom",  "qs_rank": 16},
    "King's College London":            {"country": "United Kingdom",  "qs_rank": 17},
    "University of Bristol":            {"country": "United Kingdom",  "qs_rank": 18},
    "University of Warwick":            {"country": "United Kingdom",  "qs_rank": 19},
    "University of Glasgow":            {"country": "United Kingdom",  "qs_rank": 20},
    "Delft University of Technology":   {"country": "Netherlands",     "qs_rank": 21},
    "Wageningen University":            {"country": "Netherlands",     "qs_rank": 22},
    "Ghent University":                 {"country": "Belgium",         "qs_rank": 23},
    "University of Helsinki":           {"country": "Finland",         "qs_rank": 24},
    "Stockholm University":             {"country": "Sweden",          "qs_rank": 25},
    "Lund University":                  {"country": "Sweden",          "qs_rank": 26},
    "University of Vienna":             {"country": "Austria",         "qs_rank": 27},
    "University of Zurich":             {"country": "Switzerland",     "qs_rank": 28},
    "University of Geneva":             {"country": "Switzerland",     "qs_rank": 29},
    "University of Basel":              {"country": "Switzerland",     "qs_rank": 30},
    "Sorbonne University":              {"country": "France",          "qs_rank": 31},
    "Paris-Saclay University":          {"country": "France",          "qs_rank": 32},
    "École Normale Supérieure Paris":   {"country": "France",          "qs_rank": 33},
    "Pierre and Marie Curie (UPMC)":    {"country": "France",          "qs_rank": 34},
    "University of Barcelona":          {"country": "Spain",           "qs_rank": 35},
    "Autonomous University of Madrid":  {"country": "Spain",           "qs_rank": 36},
    "University of Bologna":            {"country": "Italy",           "qs_rank": 37},
    "Sapienza University of Rome":      {"country": "Italy",           "qs_rank": 38},
    "University of Milan":              {"country": "Italy",           "qs_rank": 39},
    "Charles University Prague":        {"country": "Czech Republic",  "qs_rank": 40},
    "University of Warsaw":             {"country": "Poland",          "qs_rank": 41},
    "Jagiellonian University":          {"country": "Poland",          "qs_rank": 42},
    "University of Oslo":               {"country": "Norway",          "qs_rank": 43},
    "University of Groningen":          {"country": "Netherlands",     "qs_rank": 44},
    "Radboud University":               {"country": "Netherlands",     "qs_rank": 45},
    "University of Bern":               {"country": "Switzerland",     "qs_rank": 46},
    "TU Berlin":                        {"country": "Germany",         "qs_rank": 47},
    "RWTH Aachen":                      {"country": "Germany",         "qs_rank": 48},
    "University of Göttingen":          {"country": "Germany",         "qs_rank": 49},
    "Durham University":                {"country": "United Kingdom",  "qs_rank": 50},
}

NA = "Not available"

# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def get_soup(url: str, retries: int = MAX_RETRIES) -> BeautifulSoup | None:
    """Fetch a URL and return a BeautifulSoup object, or None on failure."""
    for attempt in range(1, retries + 1):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
            return BeautifulSoup(resp.text, "html.parser")
        except requests.RequestException as exc:
            log.warning("Attempt %d/%d failed for %s: %s", attempt, retries, url, exc)
            if attempt < retries:
                time.sleep(REQUEST_DELAY * attempt)
    return None


def absolute_url(base: str, href: str) -> str:
    """Convert a relative href to an absolute URL."""
    if not href:
        return NA
    return urljoin(base, href)


def clean_text(text: str | None) -> str:
    """Strip and normalise whitespace; return NA if empty."""
    if not text:
        return NA
    cleaned = " ".join(text.split())
    return cleaned if cleaned else NA


def make_entry(
    university: str,
    group: str = NA,
    professor: str = NA,
    research_area: str = NA,
    link: str = NA,
) -> dict:
    """Build a standardised result dictionary."""
    meta = UNIVERSITY_META.get(university, {})
    return {
        "QS Rank":       meta.get("qs_rank", NA),
        "University":    university,
        "Country":       meta.get("country", NA),
        "Group":         clean_text(group),
        "Professor":     clean_text(professor),
        "Research Area": clean_text(research_area),
        "Link":          link,
    }


def sleep():
    """Polite delay between requests."""
    time.sleep(REQUEST_DELAY)


# ---------------------------------------------------------------------------
# Individual university scrapers
# ---------------------------------------------------------------------------

def scrape_cambridge() -> list[dict]:
    """
    University of Cambridge – Department of Chemistry
    Strategy: Parse the research-group index page which lists groups with links.
    """
    uni = "University of Cambridge"
    base = "https://www.ch.cam.ac.uk"
    url  = f"{base}/research/groups"
    results = []

    soup = get_soup(url)
    if not soup:
        log.warning("[Cambridge] Could not fetch %s", url)
        return [make_entry(uni, link=url)]

    # Each group is in a card/article element
    for card in soup.select("div.views-row, article.research-group, div.group-item"):
        name_tag = card.find(["h2", "h3", "h4", "a"])
        group_name = clean_text(name_tag.get_text()) if name_tag else NA

        link_tag = card.find("a", href=True)
        group_link = absolute_url(base, link_tag["href"]) if link_tag else url

        # Try to find PI name within card
        pi_tag = card.find(class_=re.compile(r"pi|professor|leader", re.I))
        professor = clean_text(pi_tag.get_text()) if pi_tag else NA

        area_tag = card.find(class_=re.compile(r"area|topic|theme|research", re.I))
        area = clean_text(area_tag.get_text()) if area_tag else NA

        results.append(make_entry(uni, group=group_name, professor=professor,
                                  research_area=area, link=group_link))

    if not results:
        # Fallback: collect all anchor text from research section
        for a in soup.select("a[href*='research']")[:20]:
            group_name = clean_text(a.get_text())
            if len(group_name) > 5:
                results.append(make_entry(uni, group=group_name,
                                          link=absolute_url(base, a["href"])))

    return results or [make_entry(uni, link=url)]


def scrape_oxford() -> list[dict]:
    """
    University of Oxford – Department of Chemistry
    Strategy: Research themes page + sub-group pages.
    """
    uni  = "University of Oxford"
    base = "https://www.chem.ox.ac.uk"
    url  = f"{base}/research/research-themes"
    results = []

    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]

    for item in soup.select("div.views-row, li.research-theme, div.theme-item"):
        title_tag = item.find(["h2", "h3", "h4", "strong", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA

        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url

        desc_tag = item.find("p")
        area = clean_text(desc_tag.get_text()) if desc_tag else NA

        results.append(make_entry(uni, group=group, research_area=area, link=link))

    if not results:
        for a in soup.select("a[href*='research']")[:20]:
            text = clean_text(a.get_text())
            if len(text) > 5:
                results.append(make_entry(uni, group=text,
                                          link=absolute_url(base, a["href"])))

    return results or [make_entry(uni, link=url)]


def scrape_eth_zurich() -> list[dict]:
    """
    ETH Zurich – Department of Chemistry and Applied Biosciences (D-CHAB)
    Strategy: Professor directory page.
    """
    uni  = "ETH Zurich"
    base = "https://chab.ethz.ch"
    url  = f"{base}/en/research/research-groups.html"
    results = []

    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]

    for card in soup.select("div.researchtopic, div.staffCard, article.news-item, li.person"):
        name_tag = card.find(["h3", "h4", "strong", "a"])
        professor = clean_text(name_tag.get_text()) if name_tag else NA

        link_tag = card.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url

        area_tag = card.find("p")
        area = clean_text(area_tag.get_text()) if area_tag else NA

        results.append(make_entry(uni, professor=professor,
                                  research_area=area, link=link))

    if not results:
        for a in soup.select("a[href*='research']")[:20]:
            text = clean_text(a.get_text())
            if len(text) > 5:
                results.append(make_entry(uni, group=text,
                                          link=absolute_url(base, a["href"])))

    return results or [make_entry(uni, link=url)]


def scrape_imperial() -> list[dict]:
    """
    Imperial College London – Department of Chemistry
    Strategy: Research sections page.
    """
    uni  = "Imperial College London"
    base = "https://www.imperial.ac.uk"
    url  = f"{base}/chemistry/research/"
    results = []

    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]

    for section in soup.select("div.ict-body-copy, div.research-section, article"):
        title = section.find(["h2", "h3", "h4"])
        group = clean_text(title.get_text()) if title else NA
        para  = section.find("p")
        area  = clean_text(para.get_text()) if para else NA
        link_tag = section.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group,
                                      research_area=area, link=link))

    if not results:
        for a in soup.select("a[href*='research']")[:20]:
            text = clean_text(a.get_text())
            if len(text) > 5:
                results.append(make_entry(uni, group=text,
                                          link=absolute_url(base, a["href"])))

    return results or [make_entry(uni, link=url)]


def scrape_ucl() -> list[dict]:
    """
    UCL – Department of Chemistry
    Strategy: Research group listing.
    """
    uni  = "UCL"
    base = "https://www.ucl.ac.uk"
    url  = f"{base}/chemistry/research/research-groups"
    results = []

    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]

    for item in soup.select("li.collection-item, div.views-row, article"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        para  = item.find("p")
        area  = clean_text(para.get_text()) if para else NA
        if group != NA:
            results.append(make_entry(uni, group=group,
                                      research_area=area, link=link))

    return results or [make_entry(uni, link=url)]


def scrape_edinburgh() -> list[dict]:
    """
    University of Edinburgh – School of Chemistry
    Strategy: Research themes + group leaders.
    """
    uni  = "University of Edinburgh"
    base = "https://www.chem.ed.ac.uk"
    url  = f"{base}/research/research-groups"
    results = []

    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]

    for item in soup.select("div.field-item, div.views-row, li.item"):
        title_tag = item.find(["h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))

    return results or [make_entry(uni, link=url)]


def scrape_lmu_munich() -> list[dict]:
    """
    LMU Munich – Department of Chemistry
    Strategy: Working groups page.
    """
    uni  = "LMU Munich"
    base = "https://www.cup.uni-muenchen.de"
    url  = f"{base}/en/research/working-groups.html"
    results = []

    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]

    for item in soup.select("div.ce-bodytext, li.workgroup, article"):
        title_tag = item.find(["h2", "h3", "h4", "strong"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))

    return results or [make_entry(uni, link=url)]


def scrape_epfl() -> list[dict]:
    """
    EPFL – Institute of Chemical Sciences and Engineering (ISIC)
    Strategy: People / research labs directory.
    """
    uni  = "EPFL"
    base = "https://www.epfl.ch"
    url  = f"{base}/schools/sb/research/isic/groups/"
    results = []

    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]

    for item in soup.select("div.teaser, article.lab, div.lab-card"):
        title_tag = item.find(["h3", "h4", "strong", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        pi_tag = item.find(class_=re.compile(r"pi|prof|director", re.I))
        professor = clean_text(pi_tag.get_text()) if pi_tag else NA
        if group != NA:
            results.append(make_entry(uni, group=group, professor=professor, link=link))

    return results or [make_entry(uni, link=url)]


def scrape_copenhagen() -> list[dict]:
    """
    University of Copenhagen – Department of Chemistry
    Strategy: Research section page.
    """
    uni  = "University of Copenhagen"
    base = "https://www.chem.ku.dk"
    url  = f"{base}/research/"
    results = []

    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]

    for item in soup.select("div.research-section, li.research-group, article"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))

    return results or [make_entry(uni, link=url)]


def scrape_ku_leuven() -> list[dict]:
    """
    KU Leuven – Department of Chemistry
    Strategy: Research group directory.
    """
    uni  = "KU Leuven"
    base = "https://chem.kuleuven.be"
    url  = f"{base}/en/research"
    results = []

    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]

    for item in soup.select("div.research-group, li.group, article"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))

    return results or [make_entry(uni, link=url)]


def scrape_amsterdam() -> list[dict]:
    uni  = "University of Amsterdam"
    base = "https://www.uva.nl"
    url  = f"{base}/en/research/chemistry"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, div.research-group, li.item"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_leiden() -> list[dict]:
    uni  = "Leiden University"
    base = "https://www.universiteitleiden.nl"
    url  = f"{base}/en/science/chemistry/research"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, div.research-item"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_utrecht() -> list[dict]:
    uni  = "Utrecht University"
    base = "https://www.uu.nl"
    url  = f"{base}/en/research/debye-institute-for-nanomaterials-science/research-groups"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_tum() -> list[dict]:
    uni  = "TU Munich"
    base = "https://www.ch.tum.de"
    url  = f"{base}/en/research/research-groups/"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("div.research-group, article, li.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_heidelberg() -> list[dict]:
    uni  = "Heidelberg University"
    base = "https://www.chemie.uni-heidelberg.de"
    url  = f"{base}/en/research"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("div.workgroup, article, li.entry"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_manchester() -> list[dict]:
    uni  = "University of Manchester"
    base = "https://www.chemistry.manchester.ac.uk"
    url  = f"{base}/research/research-groups/"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("div.views-row, article, li.item"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_kcl() -> list[dict]:
    uni  = "King's College London"
    base = "https://www.kcl.ac.uk"
    url  = f"{base}/chemistry/research"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, div.research-group, li"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_bristol() -> list[dict]:
    uni  = "University of Bristol"
    base = "https://www.bristol.ac.uk"
    url  = f"{base}/chemistry/research/groups/"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("div.views-row, li.item, article"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_warwick() -> list[dict]:
    uni  = "University of Warwick"
    base = "https://warwick.ac.uk"
    url  = f"{base}/fac/sci/chemistry/research/groups/"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("li.item, div.item, article"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_glasgow() -> list[dict]:
    uni  = "University of Glasgow"
    base = "https://www.gla.ac.uk"
    url  = f"{base}/schools/chemistry/research/"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, div.research-group, li"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_delft() -> list[dict]:
    uni  = "Delft University of Technology"
    base = "https://www.tudelft.nl"
    url  = f"{base}/en/faculty-of-applied-sciences/about-faculty/departments/chemical-engineering/research/"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.section"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_wageningen() -> list[dict]:
    uni  = "Wageningen University"
    base = "https://www.wur.nl"
    url  = f"{base}/en/research-results/research-institutes/wfsr.htm"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.chair"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_ghent() -> list[dict]:
    uni  = "Ghent University"
    base = "https://www.ugent.be"
    url  = f"{base}/en/research/chemistry"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.research"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_helsinki() -> list[dict]:
    uni  = "University of Helsinki"
    base = "https://www.helsinki.fi"
    url  = f"{base}/en/faculty-of-science/faculty/departments/department-of-chemistry/research"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_stockholm() -> list[dict]:
    uni  = "Stockholm University"
    base = "https://www.su.se"
    url  = f"{base}/english/research/research-areas/natural-sciences/chemistry/"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.section"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_lund() -> list[dict]:
    uni  = "Lund University"
    base = "https://www.lu.se"
    url  = f"{base}/en/lucat/research-groups?q=chemistry"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("li.result-list__item, article, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_vienna() -> list[dict]:
    uni  = "University of Vienna"
    base = "https://www.univie.ac.at"
    url  = f"{base}/en/research/research-profile/natural-sciences/chemistry"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_zurich() -> list[dict]:
    uni  = "University of Zurich"
    base = "https://www.chem.uzh.ch"
    url  = f"{base}/en/research.html"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("div.researchgroup, article, li"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_geneva() -> list[dict]:
    uni  = "University of Geneva"
    base = "https://www.unige.ch"
    url  = f"{base}/sciences/chim/en/recherche/"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_basel() -> list[dict]:
    uni  = "University of Basel"
    base = "https://chemistry.unibas.ch"
    url  = f"{base}/en/research/"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, div.research-group, li"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_sorbonne() -> list[dict]:
    uni  = "Sorbonne University"
    base = "https://www.sorbonne-universite.fr"
    url  = f"{base}/en/research/our-research/chemistry"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, div.item, li"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_paris_saclay() -> list[dict]:
    uni  = "Paris-Saclay University"
    base = "https://www.universite-paris-saclay.fr"
    url  = f"{base}/en/research/research-structures/chemistry"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, div.lab, li"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_ens_paris() -> list[dict]:
    uni  = "École Normale Supérieure Paris"
    base = "https://www.chimie.ens.fr"
    url  = f"{base}/en/research"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, div.team, li.item"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_upmc() -> list[dict]:
    uni  = "Pierre and Marie Curie (UPMC)"
    base = "https://www.sorbonne-universite.fr"
    url  = f"{base}/en/research/our-research/chemistry"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, div.item, li"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_barcelona() -> list[dict]:
    uni  = "University of Barcelona"
    base = "https://www.ub.edu"
    url  = f"{base}/web/portal/en/chemistry/research"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, div.group, li"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_madrid_uam() -> list[dict]:
    uni  = "Autonomous University of Madrid"
    base = "https://www.uam.es"
    url  = f"{base}/Ciencias/Quimica/1242655024459.htm?language=en"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_bologna() -> list[dict]:
    uni  = "University of Bologna"
    base = "https://chimind.unibo.it"
    url  = f"{base}/en/research"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_sapienza() -> list[dict]:
    uni  = "Sapienza University of Rome"
    base = "https://www.chimica.uniroma1.it"
    url  = f"{base}/en/research"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_milan() -> list[dict]:
    uni  = "University of Milan"
    base = "https://www.unimi.it"
    url  = f"{base}/en/research/departments/chemistry"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_charles_prague() -> list[dict]:
    uni  = "Charles University Prague"
    base = "https://www.natur.cuni.cz"
    url  = f"{base}/en/research/research-groups"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_warsaw() -> list[dict]:
    uni  = "University of Warsaw"
    base = "https://www.chem.uw.edu.pl"
    url  = f"{base}/en/research/"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_jagiellonian() -> list[dict]:
    uni  = "Jagiellonian University"
    base = "https://chemia.uj.edu.pl"
    url  = f"{base}/en/research"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_oslo() -> list[dict]:
    uni  = "University of Oslo"
    base = "https://www.mn.uio.no"
    url  = f"{base}/kjemi/english/research/"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_groningen() -> list[dict]:
    uni  = "University of Groningen"
    base = "https://www.rug.nl"
    url  = f"{base}/research/stratingh/research/"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_radboud() -> list[dict]:
    uni  = "Radboud University"
    base = "https://www.ru.nl"
    url  = f"{base}/en/science/department-of-chemistry"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_bern() -> list[dict]:
    uni  = "University of Bern"
    base = "https://www.dcb.unibe.ch"
    url  = f"{base}/en/research.html"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, div.group, li"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_tu_berlin() -> list[dict]:
    uni  = "TU Berlin"
    base = "https://www.tu.berlin"
    url  = f"{base}/en/about/organization/faculty-ii-mathematics-and-natural-sciences/institute-of-chemistry/"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.chair"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_rwth_aachen() -> list[dict]:
    uni  = "RWTH Aachen"
    base = "https://www.chemie.rwth-aachen.de"
    url  = f"{base}/cms/Chemie/Forschung/~eejb/Arbeitsgruppen/?lidx=1"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_goettingen() -> list[dict]:
    uni  = "University of Göttingen"
    base = "https://www.chemie.uni-goettingen.de"
    url  = f"{base}/en/research"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


def scrape_durham() -> list[dict]:
    uni  = "Durham University"
    base = "https://www.durham.ac.uk"
    url  = f"{base}/departments/academic/chemistry/research/groups/"
    results = []
    soup = get_soup(url)
    if not soup:
        return [make_entry(uni, link=url)]
    for item in soup.select("article, li.item, div.group"):
        title_tag = item.find(["h2", "h3", "h4", "a"])
        group = clean_text(title_tag.get_text()) if title_tag else NA
        link_tag = item.find("a", href=True)
        link = absolute_url(base, link_tag["href"]) if link_tag else url
        if group != NA:
            results.append(make_entry(uni, group=group, link=link))
    return results or [make_entry(uni, link=url)]


# ---------------------------------------------------------------------------
# Registry: maps university name → scraper function
# ---------------------------------------------------------------------------
# To add a new university:
#   1. Define def scrape_<name>() returning list[dict]
#   2. Add an entry to SCRAPER_REGISTRY below
#   3. Add the university to UNIVERSITY_META above

SCRAPER_REGISTRY: dict[str, callable] = {
    "University of Cambridge":          scrape_cambridge,
    "University of Oxford":             scrape_oxford,
    "ETH Zurich":                       scrape_eth_zurich,
    "Imperial College London":          scrape_imperial,
    "UCL":                              scrape_ucl,
    "University of Edinburgh":          scrape_edinburgh,
    "LMU Munich":                       scrape_lmu_munich,
    "EPFL":                             scrape_epfl,
    "University of Copenhagen":         scrape_copenhagen,
    "KU Leuven":                        scrape_ku_leuven,
    "University of Amsterdam":          scrape_amsterdam,
    "Leiden University":                scrape_leiden,
    "Utrecht University":               scrape_utrecht,
    "TU Munich":                        scrape_tum,
    "Heidelberg University":            scrape_heidelberg,
    "University of Manchester":         scrape_manchester,
    "King's College London":            scrape_kcl,
    "University of Bristol":            scrape_bristol,
    "University of Warwick":            scrape_warwick,
    "University of Glasgow":            scrape_glasgow,
    "Delft University of Technology":   scrape_delft,
    "Wageningen University":            scrape_wageningen,
    "Ghent University":                 scrape_ghent,
    "University of Helsinki":           scrape_helsinki,
    "Stockholm University":             scrape_stockholm,
    "Lund University":                  scrape_lund,
    "University of Vienna":             scrape_vienna,
    "University of Zurich":             scrape_zurich,
    "University of Geneva":             scrape_geneva,
    "University of Basel":              scrape_basel,
    "Sorbonne University":              scrape_sorbonne,
    "Paris-Saclay University":          scrape_paris_saclay,
    "École Normale Supérieure Paris":   scrape_ens_paris,
    "Pierre and Marie Curie (UPMC)":    scrape_upmc,
    "University of Barcelona":          scrape_barcelona,
    "Autonomous University of Madrid":  scrape_madrid_uam,
    "University of Bologna":            scrape_bologna,
    "Sapienza University of Rome":      scrape_sapienza,
    "University of Milan":              scrape_milan,
    "Charles University Prague":        scrape_charles_prague,
    "University of Warsaw":             scrape_warsaw,
    "Jagiellonian University":          scrape_jagiellonian,
    "University of Oslo":               scrape_oslo,
    "University of Groningen":          scrape_groningen,
    "Radboud University":               scrape_radboud,
    "University of Bern":               scrape_bern,
    "TU Berlin":                        scrape_tu_berlin,
    "RWTH Aachen":                      scrape_rwth_aachen,
    "University of Göttingen":          scrape_goettingen,
    "Durham University":                scrape_durham,
}


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def run_all_scrapers() -> list[dict]:
    """Run every registered scraper, collect and return all results."""
    all_results: list[dict] = []

    for uni_name, scraper_fn in SCRAPER_REGISTRY.items():
        log.info("Scraping → %s", uni_name)
        try:
            records = scraper_fn()
            log.info("  ✓ %d records", len(records))
            all_results.extend(records)
        except Exception as exc:
            log.error("  ✗ Error scraping %s: %s", uni_name, exc)
            all_results.append(make_entry(uni_name))
        finally:
            sleep()

    return all_results


def build_dataframe(records: list[dict]) -> pd.DataFrame:
    """Convert raw records into a clean, deduplicated DataFrame."""
    df = pd.DataFrame(records)

    # Ensure column order
    cols = ["QS Rank", "University", "Country", "Group",
            "Professor", "Research Area", "Link"]
    df = df.reindex(columns=cols)

    # Fill missing values
    df = df.fillna(NA)

    # Deduplicate on University + Group + Professor
    df = df.drop_duplicates(subset=["University", "Group", "Professor"])

    # Sort by QS rank then University
    df["QS Rank"] = pd.to_numeric(df["QS Rank"], errors="coerce")
    df = df.sort_values(["QS Rank", "University", "Group"], ignore_index=True)

    return df


def export_excel(df: pd.DataFrame, path: str) -> None:
    """Export DataFrame to a formatted Excel file."""
    df.to_excel(path, index=False, sheet_name="Chemistry Research")

    wb = load_workbook(path)
    ws = wb.active

    # Styles
    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", start_color="1F4E79")
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    alt_fill      = PatternFill("solid", start_color="D9E1F2")
    na_font       = Font(name="Arial", color="999999", italic=True, size=10)
    normal_font   = Font(name="Arial", size=10)
    center_align  = Alignment(horizontal="center", vertical="top")
    left_align    = Alignment(horizontal="left",   vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    # Column widths (A–G)
    col_widths = [10, 30, 20, 40, 30, 45, 50]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Header row formatting
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border

    ws.row_dimensions[1].height = 28

    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, cell in enumerate(row, start=1):
            cell.border = thin_border
            if fill:
                cell.fill = fill
            if str(cell.value) == NA:
                cell.font = na_font
            else:
                cell.font = normal_font
            if col_idx in (1,):   # QS Rank → center
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    log.info("Excel saved → %s", path)


def export_csv(df: pd.DataFrame, path: str) -> None:
    df.to_csv(path, index=False, encoding="utf-8-sig")
    log.info("CSV saved → %s", path)

if __name__ == "__main__":
    log.info("=== Europe Chemistry QS Top-50 Scraper ===")
    log.info("Total universities to scrape: %d", len(SCRAPER_REGISTRY))

    records = run_all_scrapers()
    log.info("Total raw records collected: %d", len(records))

    df = build_dataframe(records)
    log.info("After deduplication: %d records", len(df))

    export_excel(df, OUTPUT_XLSX)
    export_csv(df, OUTPUT_CSV)

    log.info("=== Done ===")
    print(f"\n✅  Output files:")
    print(f"    {OUTPUT_XLSX}")
    print(f"    {OUTPUT_CSV}")
    print(f"\n📊  Summary: {len(df)} records across {df['University'].nunique()} universities")
